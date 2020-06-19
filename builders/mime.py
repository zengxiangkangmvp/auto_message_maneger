from .color import is_satisfy
from io import BytesIO
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.header import Header
from pandas import ExcelWriter
from os import remove as removefile
from os import environ
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill


def make_msg(title: str, content: str, mail_type="html"):
    """
    根据文字内容生成邮件消息对象

    Args:
        title: 邮件标题
        content: 邮件内容文字
        sender: 发送人
        mail_typle: 邮件主体文字的 type, 可以是 plain(纯文本) 或者 html(html邮件)

    Returns: MIMEMultipart
        一个多 MIMEtype 的邮件消息对象
    """
    msg = MIMEMultipart()
    msg.attach(MIMEText(content, mail_type, "utf-8"))
    msg["Subject"] = Header(title, "utf-8")
    return msg


def attach_figure(msg: MIMEMultipart, cid, figure):
    """
    添加以图片方式添加 matplotlib.figure
    """
    buffer = BytesIO()
    figure.savefig(buffer)
    msg = attach_image(msg, cid, buffer.getvalue())
    return msg


def attach_image(msg: MIMEMultipart, cid, data):
    """
    为消息对象添加二进制图像内容
    添加后需要在 html 文本中使用 <img src="cid:" alt="" /> 引用图片

    Args:
        msg: 消息主体
        data: 图像的二进制内容
        cid: 图像的 Content-id
    """
    img = MIMEImage(data, "PNG")
    img.add_header("Content-ID", cid)
    img.add_header("Content-Disposition", "inline")
    msg.attach(img)
    return msg


def attach_file(msg: MIMEMultipart, name: str, path: str):
    """
    为消息对象添加附件

    Args:
        msg: MIMEMultipart 类型的邮件消息对象
        name: 附件的名称
        path: 对应文件的路径

    Returns: MIMEMultipart
        附加上了附件后的 MIMEMultipart
    """

    f = open(path, "rb")
    part = MIMEApplication(f.read(), Name=name)
    f.close()
    # Content-Disposition will be like:
    # attachment; filename*=gbk''%C9%C8%B7%E7%B5%E3.xlsx
    t = ("gbk", "", name)
    part.add_header("Content-Disposition", "attachment", filename=t)
    msg.attach(part)
    return msg


def attach_df(msg: MIMEMultipart, name: str, df_set):
    """

    Args:
        df_set: df的集合, [[a,df], [b,df], [c, df]..]  a,b,c 为工作表的名称
    """
    name = _fix_excel_name(name)
    path = environ["TEMP"] + ("\\%s" % name)
    create_excel(path, df_set)
    attach_file(msg, name, path)

    try:
        removefile(path)
    except Exception as e:
        print("生成的 Excel 文件未删除,原因:\n" + repr(e))

    return msg


def _fix_excel_name(name):
    """
    调整 excel 文件名, 如果没有 .xlsx 或者 .xls 后缀, 则添加 .xlsx 后缀
    """
    if name[-5:] != ".xlsx" or name[-4:] != ".xls":
        return name + ".xlsx"
    else:
        return name


def make_fill(color):
    c = f"FF{color[1:]}"
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def create_excel(path, df_set):
    """
    根据路径和 df_set 的 工作表名称, DataFrame 创建 excel 文件
    """
    writer = ExcelWriter(path, engine="openpyxl")

    # 这里的 style 仅用于之后的表格样式套用
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )

    for item in df_set:
        sht = item[0]
        df = item[1]
        fill_rules = item[2] if len(item) > 2 else {}
        df.to_excel(writer, sheet_name=sht, index=False, engine="openpyxl")
        cols = get_col_width(df)
        for col in cols:
            col_letter = get_column_letter(col)
            col_width = cols[col]
            writer.sheets[sht].column_dimensions[col_letter].width = col_width

            # 遍历每一个单元格根据条件进行颜色填充
            col_name = writer.sheets[sht].cell(1, col).value
            rules = fill_rules.get(col_name, [])
            for i in range(2, writer.sheets[sht].max_row):
                for symbol, val, color in rules:
                    x = writer.sheets[sht].cell(i, col).value
                    if is_satisfy(x, symbol, val):
                        writer.sheets[sht].cell(i, col).fill = make_fill(color)
                        break

        # 在当前工作表创建一个 Table
        c = get_column_letter(writer.sheets[sht].max_column)
        r = writer.sheets[sht].max_row
        table = Table(displayName=sht, ref=f"A1:{c}{r}")
        table.tableStyleInfo = style
        writer.sheets[sht].add_table(table)
    writer.save()


def get_col_width(df):
    """
    根据 df 的内容计算 df 每一列在 excel 中的宽度

    Args:
        df: 计算最终列宽的 pd.DataFrame

    Returns: dict
        key 为列号, 从1开始计数, value 为列宽, 即 excel 中看到的列宽
        类似: {1:10, 2:32, 3:15 ...}
    """
    result = {}
    cols = df.columns.to_list()
    for i in range(0, len(cols)):
        title_width = get_width(cols[i])
        content_width = df[cols[i]].astype("str").apply(lambda x: get_width(x))
        content_width = content_width.max()
        width = max(content_width, title_width)
        width = 9 if width < 9 else width
        width = 50 if width > 50 else width
        result[i + 1] = width + 1
    return result


def get_width(text):
    """
    获取 text 文本的长度

    Returns: int
        text 中 ascii字符数量 * 1 + 非ascii字符数量 * 2
    """
    count_ascii = len(list(filter(lambda x: x.isascii(), text)))
    count_other = len(text) - count_ascii
    result = count_ascii + count_other * 2
    return result
